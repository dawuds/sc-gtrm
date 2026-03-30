#!/usr/bin/env python3
"""
Generate SC-GTRM Audit Work Program (AWP) Excel workbook.

Reads JSON data from awp-data/ and produces a professional AWP workbook
matching the DAC-Cybersecurity-Custody-WorkProgram.xlsx format.

Usage: python3 workprogram/generate-awp.py
Output: workprogram/SC-GTRM-AWP.xlsx
"""

import json
import os
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, NamedStyle
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("openpyxl required: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).parent
AWP_DATA_DIR = SCRIPT_DIR / "awp-data"
OUTPUT_FILE = SCRIPT_DIR / "SC-GTRM-AWP.xlsx"

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
FONT_SANS = "Calibri"

# Colors
NAVY = "1F3864"
DARK_BLUE = "2B4C7E"
MID_BLUE = "4472C4"
LIGHT_BLUE = "D6E4F0"
PALE_BLUE = "E9EFF7"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"
DARK_GREY = "404040"
MID_GREY = "808080"
AMBER_BG = "FFF2CC"
AMBER_BORDER = "BF9000"
GREEN_BG = "E2EFDA"
GREEN_TEXT = "375623"
RED_BG = "FCE4EC"
RED_TEXT = "C62828"
ORANGE_BG = "FFF3E0"
ORANGE_TEXT = "E65100"
GREY_BG = "ECEFF1"
GREY_TEXT = "546E7A"
CONTEXT_BG = "EBF0F7"
CONTEXT_BORDER = "8EAADB"
GOLD = "DAA520"
SCORE_GREEN_BG = "C6EFCE"
SCORE_GREEN_TEXT = "006100"
SCORE_AMBER_BG = "FFEB9C"
SCORE_AMBER_TEXT = "9C5700"
SCORE_RED_BG = "FFC7CE"
SCORE_RED_TEXT = "9C0006"
CRITICAL_BG = "F8D7DA"
CRITICAL_TEXT = "842029"
CONDITIONAL_BG = "FFF3CD"
CONDITIONAL_TEXT = "664D03"
STANDARD_BG = "D1E7DD"
STANDARD_TEXT = "0F5132"

# Borders
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)
BOTTOM_BORDER = Border(bottom=Side(style="thin", color="B4B4B4"))
THICK_BOTTOM = Border(bottom=Side(style="medium", color=NAVY))

# Fills
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
SECTION_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color=PALE_BLUE, end_color=PALE_BLUE, fill_type="solid")
LIGHT_FILL = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
AMBER_FILL = PatternFill(start_color=AMBER_BG, end_color=AMBER_BG, fill_type="solid")
METHODOLOGY_HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")

# Fonts
TITLE_FONT = Font(name=FONT_SANS, bold=True, size=14, color=NAVY)
SUBTITLE_FONT = Font(name=FONT_SANS, bold=False, size=11, color=MID_GREY)
SECTION_TITLE_FONT = Font(name=FONT_SANS, bold=True, size=11, color=NAVY)
HEADER_FONT = Font(name=FONT_SANS, bold=True, size=10, color=WHITE)
SECTION_HEADER_FONT = Font(name=FONT_SANS, bold=True, size=10, color=WHITE)
BODY_FONT = Font(name=FONT_SANS, size=10, color=DARK_GREY)
BODY_BOLD = Font(name=FONT_SANS, bold=True, size=10, color=DARK_GREY)
SMALL_FONT = Font(name=FONT_SANS, size=9, color=MID_GREY)
METH_LABEL_FONT = Font(name=FONT_SANS, bold=True, size=10, color=NAVY)
METH_VALUE_FONT = Font(name=FONT_SANS, size=10, color=DARK_GREY)
METH_HEADER_FONT = Font(name=FONT_SANS, bold=True, size=11, color=WHITE)
DISCLAIMER_FONT = Font(name=FONT_SANS, italic=True, size=9, color="92400E")
CONTEXT_FONT = Font(name=FONT_SANS, italic=True, size=9, color=DARK_BLUE)
CONTEXT_BOLD = Font(name=FONT_SANS, bold=True, italic=True, size=9, color=NAVY)
TIER_FONT = Font(name=FONT_SANS, bold=True, size=9)

# Alignments
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_TOP = Alignment(vertical="top")

# ---------------------------------------------------------------------------
# Assessment sheet column config
# ---------------------------------------------------------------------------
ASSESSMENT_COLUMNS = [
    ("Ref", 8),
    ("Control", 22),
    ("Sub-Procedure", 24),
    ("Assessment Procedures", 48),
    ("Expected Evidence", 36),
    ("Method", 14),
    ("Procedures Performed", 30),
    ("Evidence Obtained", 24),
    ("Evidence Ref", 12),
    ("Observation / Findings", 30),
    ("Conclusion", 14),
    ("Recommendations", 30),
]


def load_sheet_data():
    """Load all AWP sheet JSON files."""
    sheets = []
    for fname in sorted(AWP_DATA_DIR.glob("sheet*.json")):
        with open(fname) as f:
            sheets.append(json.load(f))
    return sheets


# ---------------------------------------------------------------------------
# Sheet 1: Methodology & Approach
# ---------------------------------------------------------------------------
def build_methodology_sheet(wb):
    ws = wb.active
    ws.title = "Methodology & Approach"
    ws.sheet_properties.tabColor = NAVY

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50
    for col in "CDEFG":
        ws.column_dimensions[col].width = 14

    row = 1

    # --- Title ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1,
                   value="INDEPENDENT ASSESSMENT \u2014 METHODOLOGY & APPROACH")
    cell.font = Font(name=FONT_SANS, bold=True, size=14, color=WHITE)
    cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 36
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1,
                   value="SC-GL/6-2023 Guidelines on Technology Risk Management \u2014 Capital Market Intermediary")
    cell.font = Font(name=FONT_SANS, size=11, color=WHITE)
    cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24
    row += 2

    # --- Engagement Details ---
    def section_header(title):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = SECTION_TITLE_FONT
        cell.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        cell.border = THICK_BOTTOM
        ws.row_dimensions[row].height = 22
        row += 1

    def field_row(label, value):
        nonlocal row
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = METH_LABEL_FONT
        c1.alignment = LEFT_TOP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        c2 = ws.cell(row=row, column=2, value=value)
        c2.font = METH_VALUE_FONT
        c2.alignment = LEFT_TOP
        c2.border = BOTTOM_BORDER
        row += 1

    section_header("ENGAGEMENT DETAILS")
    field_row("Entity Name:", "[Entity Name]")
    field_row("Regulated Activity:", "Capital Market Intermediary (CMSL/CMSRL Holder)")
    field_row("Assessment Date:", "[DD/MM/YYYY]")
    field_row("Assessment Period:", "[Start Date] to [End Date]")
    field_row("Lead Assessor:", "[Name, Qualification]")
    field_row("Engagement Reference:", "[Ref Number]")
    field_row("Report Classification:", "Confidential (Sulit)")
    row += 1

    # --- Scope ---
    section_header("SCOPE OF ASSESSMENT")
    scope_lines = [
        "This assessment is conducted pursuant to SC-GL/6-2023 Guidelines on Technology Risk Management for capital market intermediaries regulated by the Securities Commission Malaysia.",
        "The objective is to ascertain that the entity has established and maintains an effective technology risk management framework covering governance, cybersecurity, resilience, third-party management, data protection, and emerging technology risks.",
        "Assessment covers 35 controls across 10 domains: (1) Governance & Oversight, (2) Technology Risk Assessment, (3) Cybersecurity, (4) Business Continuity, (5) Incident Management, (6) Third-Party Risk, (7) Cloud Services, (8) Data Management, (9) Technology Audit, (10) Emerging Technology.",
        "IMPORTANT: Third-party validation does not guarantee SC approval. Areas assessed reflect the assessor's professional judgement guided by SC-GL/6-2023 and may not cover every SC expectation.",
    ]
    for line in scope_lines:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value=line)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        ws.row_dimensions[row].height = 40
        row += 1
    row += 1

    # --- Assessment Methods ---
    section_header("ASSESSMENT METHODS")
    methods = [
        ("Inspection", "Examination of documents, records, policies, configurations, and tangible evidence to verify existence, completeness, and compliance with SC-GL/6-2023 requirements."),
        ("Inquiry", "Interviews with management, process owners, and staff to understand controls and their operating effectiveness within the capital market entity."),
        ("Observation", "Direct observation of processes, systems, and activities being performed."),
        ("Confirmation", "Verification with external or independent parties (e.g., CSP certifications, third-party audit reports)."),
    ]
    for label, desc in methods:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = BODY_BOLD
        c1.alignment = LEFT_TOP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        c2 = ws.cell(row=row, column=2, value=desc)
        c2.font = BODY_FONT
        c2.alignment = WRAP
        ws.row_dimensions[row].height = 36
        row += 1
    row += 1

    # --- Sampling Methodology ---
    section_header("SAMPLING METHODOLOGY")

    # Approach paragraph
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1,
                   value="This assessment uses judgemental (non-statistical) sampling. "
                         "Sample sizes are determined by population size and control risk tier. "
                         "The assessor exercises professional judgement in sample selection, "
                         "prioritising items with higher inherent risk, recent changes, "
                         "or anomalies identified during the assessment.")
    cell.font = BODY_FONT
    cell.alignment = WRAP
    ws.row_dimensions[row].height = 48
    row += 2

    # Table 1: Population-based sample sizes
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value="Table 1 \u2014 Population-Based Sample Sizes")
    cell.font = METH_LABEL_FONT
    row += 1

    sample_headers = ["Population Size", "Standard Risk", "High Risk (Critical Tier)", "Rationale"]
    for ci, hdr in enumerate(sample_headers, 1):
        c = ws.cell(row=row, column=ci, value=hdr)
        c.font = Font(name=FONT_SANS, bold=True, size=9, color=WHITE)
        c.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
        c.alignment = CENTER
        c.border = THIN_BORDER
    row += 1

    sample_rows = [
        ("1\u20135", "All", "All", "Full coverage feasible"),
        ("6\u201315", "3", "5", "~25\u201340% coverage"),
        ("16\u201350", "5", "8", "Sufficient for pattern detection"),
        ("51\u2013100", "8", "12", "Diminishing returns beyond this"),
        ("100+", "10", "15", "Cap with stratification"),
    ]
    for vals in sample_rows:
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = BODY_FONT
            c.alignment = CENTER
            c.border = THIN_BORDER
        row += 1
    row += 1

    # Table 2: Time-based evidence coverage
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value="Table 2 \u2014 Time-Based Evidence Coverage")
    cell.font = METH_LABEL_FONT
    row += 1

    time_headers = ["Evidence Type", "Coverage Period", "Rationale", ""]
    for ci, hdr in enumerate(time_headers, 1):
        c = ws.cell(row=row, column=ci, value=hdr if hdr else None)
        c.font = Font(name=FONT_SANS, bold=True, size=9, color=WHITE)
        c.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
        c.alignment = CENTER
        c.border = THIN_BORDER
    row += 1

    time_rows = [
        ("Board / committee reports", "Last 4 quarters", "Full annual governance cycle"),
        ("Operational records (incidents, changes, patches)", "Last 3\u20136 months", "Recent operating effectiveness"),
        ("Annual processes (risk assessment, DR test, pen test)", "Last 12 months", "Full cycle"),
        ("Continuous monitoring (SOC, logs, alerts)", "Last 30 days", "Current state verification"),
        ("Policies and frameworks", "Current version + prior", "Change tracking"),
    ]
    for vals in time_rows:
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = BODY_FONT
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = THIN_BORDER
        ws.row_dimensions[row].height = 28
        row += 1
    row += 1

    # Table 3: Risk tier classification
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value="Table 3 \u2014 Control Risk Tier Classification")
    cell.font = METH_LABEL_FONT
    row += 1

    tier_headers = ["Risk Tier", "Sampling Level", "Description", "Controls"]
    for ci, hdr in enumerate(tier_headers, 1):
        c = ws.cell(row=row, column=ci, value=hdr)
        c.font = Font(name=FONT_SANS, bold=True, size=9, color=WHITE)
        c.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
        c.alignment = CENTER
        c.border = THIN_BORDER
    row += 1

    tier_rows = [
        ("Critical", "High Risk column",
         "Foundation controls whose failure directly impacts market operations, regulatory standing, or client protection",
         "GRA-1, CYB-2, RIM-1, RIM-2, RIM-5, RIM-6, DET-3"),
        ("Standard", "Standard Risk column",
         "Controls that support effective technology risk management but whose absence has indirect or contained impact",
         "GRA-2\u201310, CYB-1/3/4, RIM-3/4/7/8, TPC-1\u20137, DET-1/2"),
        ("Conditional", "Standard Risk (if applicable)",
         "Controls applicable only if the entity engages in specific activities. Mark N/A with justification if not applicable.",
         "DET-4, DET-5, DET-6"),
    ]
    for vals in tier_rows:
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = BODY_FONT
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = THIN_BORDER
            # Color the tier label
            if ci == 1:
                if val == "Critical":
                    c.fill = PatternFill(start_color=CRITICAL_BG, end_color=CRITICAL_BG, fill_type="solid")
                    c.font = Font(name=FONT_SANS, bold=True, size=10, color=CRITICAL_TEXT)
                elif val == "Standard":
                    c.fill = PatternFill(start_color=STANDARD_BG, end_color=STANDARD_BG, fill_type="solid")
                    c.font = Font(name=FONT_SANS, bold=True, size=10, color=STANDARD_TEXT)
                elif val == "Conditional":
                    c.fill = PatternFill(start_color=CONDITIONAL_BG, end_color=CONDITIONAL_BG, fill_type="solid")
                    c.font = Font(name=FONT_SANS, bold=True, size=10, color=CONDITIONAL_TEXT)
        ws.row_dimensions[row].height = 48
        row += 1
    row += 1

    # --- Conclusion Scale ---
    section_header("CONCLUSION SCALE")
    conclusions = [
        ("Compliant", "Control/requirement fully implemented and operating effectively. Evidence supports compliance with SC-GL/6-2023.", GREEN_BG, GREEN_TEXT),
        ("Partially Compliant", "Implemented but with gaps in scope, coverage, documentation, or effectiveness. Specific improvements required.", ORANGE_BG, ORANGE_TEXT),
        ("Non-Compliant", "Absent, fundamentally deficient, or not operating. SC-GL/6-2023 criteria not met.", RED_BG, RED_TEXT),
        ("N/A", "Not applicable to entity's operations. Justification documented.", GREY_BG, GREY_TEXT),
    ]
    for label, desc, bg, fg in conclusions:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(name=FONT_SANS, bold=True, size=10, color=fg)
        c1.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        c1.alignment = LEFT_TOP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        c2 = ws.cell(row=row, column=2, value=desc)
        c2.font = BODY_FONT
        c2.alignment = WRAP
        ws.row_dimensions[row].height = 30
        row += 1
    row += 1

    # --- Limitations ---
    section_header("LIMITATIONS")
    limitations = [
        "1. Limited assurance engagement \u2014 conclusions based on procedures performed, not an audit opinion.",
        "2. Point-in-time assessment. Does not guarantee continued compliance.",
        "3. Reliance on management representations and documentation provided.",
        "4. Scope guided by SC-GL/6-2023 Guidelines on Technology Risk Management. SC may update requirements.",
        "5. Does not constitute legal or regulatory advice.",
    ]
    for lim in limitations:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value=lim)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        ws.row_dimensions[row].height = 20
        row += 1
    row += 2

    # --- Sign-off ---
    section_header("SIGN-OFF")
    signoffs = [
        ("Lead Assessor:", "________________________", "Date:", "______________"),
        ("Quality Reviewer:", "________________________", "Date:", "______________"),
        ("Entity Representative:", "________________________", "Date:", "______________"),
    ]
    for label, sig, dlabel, dsig in signoffs:
        ws.cell(row=row, column=1, value=label).font = METH_LABEL_FONT
        ws.cell(row=row, column=2, value=sig).font = METH_VALUE_FONT
        ws.cell(row=row, column=2).border = BOTTOM_BORDER
        ws.cell(row=row, column=5, value=dlabel).font = METH_LABEL_FONT
        ws.cell(row=row, column=6, value=dsig).font = METH_VALUE_FONT
        ws.cell(row=row, column=6).border = BOTTOM_BORDER
        ws.row_dimensions[row].height = 24
        row += 1

    # Freeze top rows
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Assessment sheets (sheets 2-6)
# ---------------------------------------------------------------------------
def build_assessment_sheet(wb, sheet_data):
    """Build an assessment sheet and return control scoring metadata.

    Returns a list of dicts, one per control (section):
        {
            "ref": "GRA-1",
            "controlName": "Technology Risk Governance Framework",
            "domain": "Governance & Oversight",
            "riskTier": "critical",
            "sheetName": "Governance, Risk & Audit",
            "scoreCells": ["M3", "M4", "M5", ...]
        }
    """
    sheet_title = sheet_data["sheetName"][:31]  # Excel 31-char limit
    ws = wb.create_sheet(title=sheet_title)

    # Tab color by sheet number
    tab_colors = {2: "4472C4", 3: "C0392B", 4: "E67E22", 5: "27AE60", 6: "8E44AD"}
    ws.sheet_properties.tabColor = tab_colors.get(sheet_data["sheetNumber"], MID_BLUE)

    # Column widths (A-L assessment columns)
    for idx, (_, width) in enumerate(ASSESSMENT_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Hidden scoring helper column M
    ws.column_dimensions["M"].hidden = True

    # Data validation for Conclusion column (K)
    conclusion_dv = DataValidation(
        type="list",
        formula1='"Compliant,Partially Compliant,Non-Compliant,N/A"',
        allow_blank=True,
    )
    conclusion_dv.error = "Please select a valid conclusion"
    conclusion_dv.errorTitle = "Invalid Conclusion"
    conclusion_dv.prompt = "Select the assessment conclusion"
    conclusion_dv.promptTitle = "Conclusion"
    ws.add_data_validation(conclusion_dv)

    row = 1

    # --- Header row ---
    for col_idx, (col_name, _) in enumerate(ASSESSMENT_COLUMNS, 1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_CENTER
        cell.border = THIN_BORDER
    # Header for hidden score column
    score_hdr = ws.cell(row=row, column=13, value="Score")
    score_hdr.font = HEADER_FONT
    score_hdr.fill = HEADER_FILL
    score_hdr.alignment = WRAP_CENTER
    ws.row_dimensions[row].height = 28
    row += 1

    # --- Data rows ---
    control_scores = []  # return value
    proc_idx = 0
    for section in sheet_data["sections"]:
        # Risk tier badge for section header
        risk_tier = section.get("riskTier", "standard")
        tier_label = {"critical": " [CRITICAL]", "conditional": " [CONDITIONAL]", "standard": ""}
        header_text = section["sectionHeader"] + tier_label.get(risk_tier, "")

        # Section header row (merged across all 12 columns)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        cell = ws.cell(row=row, column=1, value=header_text)
        cell.font = SECTION_HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 24
        # Color section header by risk tier
        if risk_tier == "critical":
            cell.fill = PatternFill(start_color="8B1A1A", end_color="8B1A1A", fill_type="solid")
        elif risk_tier == "conditional":
            cell.fill = PatternFill(start_color="7D6608", end_color="7D6608", fill_type="solid")
        else:
            cell.fill = SECTION_FILL
        row += 1

        # Control context row (purpose, expectation, key risk)
        obj = section.get("controlObjective")
        if obj:
            context_text = (
                f"PURPOSE: {obj.get('purpose', '')}\n"
                f"EXPECTATION: {obj.get('expectation', '')}\n"
                f"KEY RISK: {obj.get('keyRisk', '')}"
            )
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
            cell = ws.cell(row=row, column=1, value=context_text)
            cell.font = CONTEXT_FONT
            cell.fill = PatternFill(start_color=CONTEXT_BG, end_color=CONTEXT_BG, fill_type="solid")
            cell.alignment = WRAP
            cell.border = Border(
                left=Side(style="thin", color=CONTEXT_BORDER),
                right=Side(style="thin", color=CONTEXT_BORDER),
                top=Side(style="thin", color=CONTEXT_BORDER),
                bottom=Side(style="thin", color=CONTEXT_BORDER),
            )
            ws.row_dimensions[row].height = 72
            row += 1

        # Track score cells for this control
        section_score_cells = []

        # Sub-procedures
        for sp in section["subProcedures"]:
            is_alt = proc_idx % 2 == 1
            fill = ALT_ROW_FILL if is_alt else PatternFill(fill_type=None)

            values = [
                section["ref"],
                section["controlName"],
                sp["subProcedure"],
                sp["assessmentProcedures"],
                sp["expectedEvidence"],
                sp["method"],
                "",  # Procedures Performed
                "",  # Evidence Obtained
                "",  # Evidence Ref
                "",  # Observation / Findings
                "",  # Conclusion
                "",  # Recommendations
            ]

            # Calculate row height based on content length
            max_len = max(len(str(v)) for v in values[:6])
            row_height = max(45, min(120, max_len // 2))

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = BODY_FONT
                cell.alignment = WRAP
                cell.border = THIN_BORDER
                if is_alt:
                    cell.fill = fill

                # Bold the ref column
                if col_idx == 1:
                    cell.font = BODY_BOLD
                    cell.alignment = WRAP_CENTER

                # Center the method column
                if col_idx == 6:
                    cell.alignment = WRAP_CENTER

                # Light yellow for blank working columns (7-12)
                if col_idx >= 7:
                    cell.fill = PatternFill(
                        start_color="FFFFF5", end_color="FFFFF5", fill_type="solid"
                    )

            # Add data validation to Conclusion cell (column K = 11)
            conclusion_dv.add(ws.cell(row=row, column=11))

            # Hidden scoring formula in column M (13)
            score_formula = (
                f'=IF(K{row}="Compliant",3,'
                f'IF(K{row}="Partially Compliant",2,'
                f'IF(K{row}="Non-Compliant",1,"")))'
            )
            ws.cell(row=row, column=13, value=score_formula)

            section_score_cells.append(f"M{row}")

            ws.row_dimensions[row].height = row_height
            row += 1
            proc_idx += 1

        # Record control scoring metadata
        control_scores.append({
            "ref": section["ref"],
            "controlName": section["controlName"],
            "domain": section["domain"],
            "riskTier": section.get("riskTier", "standard"),
            "sheetName": sheet_title,
            "scoreCells": section_score_cells,
        })

    # --- Disclaimer row ---
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(
        row=row, column=1,
        value="This work program is constructed for educational purposes. "
              "Customise for your engagement. Not legal or regulatory advice. "
              "AI-generated content \u2014 review and validate before use."
    )
    cell.font = DISCLAIMER_FONT
    cell.fill = AMBER_FILL
    cell.alignment = WRAP
    ws.row_dimensions[row].height = 28

    # Freeze header + panes
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(12)}1"

    return control_scores


# ---------------------------------------------------------------------------
# Scoring Summary sheet
# ---------------------------------------------------------------------------
def _quote_sheet(name):
    """Return a single-quoted sheet name safe for use in Excel formulas."""
    return f"'{name}'"


def _score_cell_range(score_cells):
    """Extract start and end row numbers from a list like ['M3','M4','M5']."""
    rows = [int(c[1:]) for c in score_cells]
    return min(rows), max(rows)


def build_scoring_sheet(wb, all_control_scores):
    """Build the Scoring Summary sheet with control, domain, and overall scores."""
    ws = wb.create_sheet(title="Scoring Summary")
    ws.sheet_properties.tabColor = GOLD

    # Column widths for Section A
    col_widths = {
        "A": 10, "B": 36, "C": 24, "D": 12, "E": 15,
        "F": 12, "G": 12, "H": 15, "I": 8, "J": 10, "K": 18,
    }
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = w

    row = 1

    # ======================================================================
    # Section A: Control-Level Scoring Table
    # ======================================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    cell = ws.cell(row=row, column=1,
                   value="SECTION A \u2014 CONTROL-LEVEL SCORING")
    cell.font = Font(name=FONT_SANS, bold=True, size=12, color=WHITE)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 1

    # Headers
    section_a_headers = [
        "Ref", "Control", "Domain", "Risk Tier", "Sub-Procedures",
        "Compliant", "Partial", "Non-Compliant", "N/A", "Score", "Conclusion",
    ]
    for ci, hdr in enumerate(section_a_headers, 1):
        c = ws.cell(row=row, column=ci, value=hdr)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = WRAP_CENTER
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 28
    header_row_a = row
    row += 1

    # Track rows per domain for Section B aggregation
    # domain -> list of row numbers in Section A
    domain_rows = {}
    control_data_start_row = row

    for ctrl in all_control_scores:
        qs = _quote_sheet(ctrl["sheetName"])
        start_r, end_r = _score_cell_range(ctrl["scoreCells"])
        k_range = f"{qs}!K{start_r}:K{end_r}"
        m_range = f"{qs}!M{start_r}:M{end_r}"

        # A: Ref
        c = ws.cell(row=row, column=1, value=ctrl["ref"])
        c.font = BODY_BOLD
        c.alignment = WRAP_CENTER
        c.border = THIN_BORDER

        # B: Control name
        c = ws.cell(row=row, column=2, value=ctrl["controlName"])
        c.font = BODY_FONT
        c.alignment = WRAP
        c.border = THIN_BORDER

        # C: Domain
        c = ws.cell(row=row, column=3, value=ctrl["domain"])
        c.font = BODY_FONT
        c.alignment = WRAP
        c.border = THIN_BORDER

        # D: Risk Tier
        tier = ctrl["riskTier"]
        c = ws.cell(row=row, column=4, value=tier.capitalize())
        c.font = Font(name=FONT_SANS, bold=True, size=9)
        c.alignment = CENTER
        c.border = THIN_BORDER
        if tier == "critical":
            c.fill = PatternFill(start_color=CRITICAL_BG, end_color=CRITICAL_BG, fill_type="solid")
            c.font = Font(name=FONT_SANS, bold=True, size=9, color=CRITICAL_TEXT)
        elif tier == "conditional":
            c.fill = PatternFill(start_color=CONDITIONAL_BG, end_color=CONDITIONAL_BG, fill_type="solid")
            c.font = Font(name=FONT_SANS, bold=True, size=9, color=CONDITIONAL_TEXT)
        else:
            c.fill = PatternFill(start_color=STANDARD_BG, end_color=STANDARD_BG, fill_type="solid")
            c.font = Font(name=FONT_SANS, bold=True, size=9, color=STANDARD_TEXT)

        # E: Sub-Procedures (static count)
        c = ws.cell(row=row, column=5, value=len(ctrl["scoreCells"]))
        c.font = BODY_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER

        # F: Compliant count
        c = ws.cell(row=row, column=6,
                     value=f'=COUNTIF({k_range},"Compliant")')
        c.font = BODY_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER

        # G: Partial count
        c = ws.cell(row=row, column=7,
                     value=f'=COUNTIF({k_range},"Partially Compliant")')
        c.font = BODY_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER

        # H: Non-Compliant count
        c = ws.cell(row=row, column=8,
                     value=f'=COUNTIF({k_range},"Non-Compliant")')
        c.font = BODY_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER

        # I: N/A count
        c = ws.cell(row=row, column=9,
                     value=f'=COUNTIF({k_range},"N/A")')
        c.font = BODY_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER

        # J: Score (average of numeric helper column, excludes blanks)
        c = ws.cell(row=row, column=10,
                     value=f'=IFERROR(AVERAGE({m_range}),"")')
        c.font = Font(name=FONT_SANS, bold=True, size=10, color=DARK_GREY)
        c.alignment = CENTER
        c.border = THIN_BORDER
        c.number_format = "0.00"

        # K: Conclusion (text based on score)
        c = ws.cell(row=row, column=11,
                     value=f'=IF(J{row}="","",IF(J{row}>=2.5,"Compliant",'
                           f'IF(J{row}>=1.5,"Partially Compliant","Non-Compliant")))')
        c.font = BODY_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER

        # Track domain membership
        domain = ctrl["domain"]
        domain_rows.setdefault(domain, []).append(row)

        ws.row_dimensions[row].height = 22
        row += 1

    control_data_end_row = row - 1

    # Add blank separator
    row += 2

    # ======================================================================
    # Section B: Domain Summary Table
    # ======================================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1,
                   value="SECTION B \u2014 DOMAIN SUMMARY")
    cell.font = Font(name=FONT_SANS, bold=True, size=12, color=WHITE)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 1

    domain_headers = [
        "Domain", "Controls", "Avg Score", "Compliant", "Partial",
        "Non-Compliant", "Overall",
    ]
    for ci, hdr in enumerate(domain_headers, 1):
        c = ws.cell(row=row, column=ci, value=hdr)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = WRAP_CENTER
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 28
    row += 1

    # Canonical domain order
    domain_order = [
        "Governance & Oversight",
        "Technology Risk Assessment",
        "Cybersecurity",
        "Business Continuity",
        "Incident Management",
        "Third-Party Risk",
        "Cloud Services",
        "Data Management",
        "Technology Audit",
        "Emerging Technology",
    ]

    domain_score_rows = {}  # domain -> row in Section B (for Section C)

    for domain in domain_order:
        d_rows = domain_rows.get(domain, [])
        if not d_rows:
            continue

        # A: Domain name
        c = ws.cell(row=row, column=1, value=domain)
        c.font = BODY_BOLD
        c.alignment = WRAP
        c.border = THIN_BORDER

        # B: Controls count (static)
        c = ws.cell(row=row, column=2, value=len(d_rows))
        c.font = BODY_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER

        # C: Avg Score — AVERAGE of the Score column (J) for this domain's rows
        score_refs = ",".join(f"J{r}" for r in d_rows)
        c = ws.cell(row=row, column=3,
                     value=f'=IFERROR(AVERAGE({score_refs}),"")')
        c.font = Font(name=FONT_SANS, bold=True, size=10, color=DARK_GREY)
        c.alignment = CENTER
        c.border = THIN_BORDER
        c.number_format = "0.00"

        # D: Compliant — SUM of Compliant counts for this domain
        comp_refs = ",".join(f"F{r}" for r in d_rows)
        c = ws.cell(row=row, column=4, value=f"=SUM({comp_refs})")
        c.font = BODY_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER

        # E: Partial
        part_refs = ",".join(f"G{r}" for r in d_rows)
        c = ws.cell(row=row, column=5, value=f"=SUM({part_refs})")
        c.font = BODY_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER

        # F: Non-Compliant
        nc_refs = ",".join(f"H{r}" for r in d_rows)
        c = ws.cell(row=row, column=6, value=f"=SUM({nc_refs})")
        c.font = BODY_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER

        # G: Overall conclusion
        c = ws.cell(row=row, column=7,
                     value=f'=IF(C{row}="","",IF(C{row}>=2.5,"Compliant",'
                           f'IF(C{row}>=1.5,"Partially Compliant","Non-Compliant")))')
        c.font = BODY_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER

        domain_score_rows[domain] = row
        ws.row_dimensions[row].height = 22
        row += 1

    domain_section_end = row - 1

    # Add blank separator
    row += 2

    # ======================================================================
    # Section C: Overall Assessment
    # ======================================================================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1,
                   value="SECTION C \u2014 OVERALL ASSESSMENT")
    cell.font = Font(name=FONT_SANS, bold=True, size=12, color=WHITE)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 1

    # All control score cells (J column) and conclusion cells (K column)
    all_j = ",".join(f"J{r}" for rows in domain_rows.values() for r in rows)
    all_k_range = f"K{control_data_start_row}:K{control_data_end_row}"

    # Row: Total controls assessed
    c = ws.cell(row=row, column=1, value="Total Controls")
    c.font = METH_LABEL_FONT
    c.border = THIN_BORDER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2, value=len(all_control_scores))
    c.font = BODY_FONT
    c.alignment = CENTER
    c.border = THIN_BORDER
    row += 1

    # Row: Controls assessed (excluding N/A conclusion)
    c = ws.cell(row=row, column=1, value="Controls Assessed (excl. N/A)")
    c.font = METH_LABEL_FONT
    c.border = THIN_BORDER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2,
                 value=f'={len(all_control_scores)}-COUNTIF({all_k_range},"N/A")')
    c.font = BODY_FONT
    c.alignment = CENTER
    c.border = THIN_BORDER
    assessed_row = row
    row += 1

    # Row: Overall score (average of domain avg scores)
    domain_avg_refs = ",".join(f"C{r}" for r in domain_score_rows.values())
    c = ws.cell(row=row, column=1, value="Overall Score")
    c.font = METH_LABEL_FONT
    c.border = THIN_BORDER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2,
                 value=f'=IFERROR(AVERAGE({domain_avg_refs}),"")')
    c.font = Font(name=FONT_SANS, bold=True, size=12, color=DARK_GREY)
    c.alignment = CENTER
    c.border = THIN_BORDER
    c.number_format = "0.00"
    overall_score_row = row
    row += 1

    # Row: Overall conclusion
    c = ws.cell(row=row, column=1, value="Overall Conclusion")
    c.font = METH_LABEL_FONT
    c.border = THIN_BORDER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2,
                 value=f'=IF(B{overall_score_row}="","",IF(B{overall_score_row}>=2.5,'
                       f'"Compliant",IF(B{overall_score_row}>=1.5,'
                       f'"Partially Compliant","Non-Compliant")))')
    c.font = Font(name=FONT_SANS, bold=True, size=12, color=DARK_GREY)
    c.alignment = CENTER
    c.border = THIN_BORDER
    row += 1

    # Row: Compliance percentage
    # Compliant count from all controls / (total - N/A)
    all_f = f"F{control_data_start_row}:F{control_data_end_row}"
    c = ws.cell(row=row, column=1, value="Compliance %")
    c.font = METH_LABEL_FONT
    c.border = THIN_BORDER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2,
                 value=f'=IFERROR(SUM({all_f})/B{assessed_row}*100,"")')
    c.font = Font(name=FONT_SANS, bold=True, size=12, color=DARK_GREY)
    c.alignment = CENTER
    c.border = THIN_BORDER
    c.number_format = "0.0"
    row += 1

    # ======================================================================
    # Apply score color formatting to Section A Score column (J)
    # ======================================================================
    for r in range(control_data_start_row, control_data_end_row + 1):
        _apply_score_formatting(ws, r, 10)  # column J

    # Apply score color formatting to Section B Avg Score column (C)
    for r in domain_score_rows.values():
        _apply_score_formatting(ws, r, 3)  # column C

    # Apply score formatting to overall score
    _apply_score_formatting(ws, overall_score_row, 2)  # column B

    # Freeze first row
    ws.freeze_panes = "A2"


def _apply_score_formatting(ws, row, col):
    """Apply conditional-style formatting to a score cell.

    Since openpyxl conditional formatting with formulas can be complex,
    we use a neutral format here. The actual color will show once the
    workbook is opened and values are calculated. For a static approach,
    we leave formatting neutral and let the conclusion text convey status.

    However, we CAN set number format and a subtle style.
    """
    # The cell already has a formula; we just ensure number format is set.
    c = ws.cell(row=row, column=col)
    c.number_format = "0.00"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Loading AWP data...")
    sheets = load_sheet_data()
    print(f"  Loaded {len(sheets)} assessment sheets")

    total_subs = sum(
        len(sp)
        for s in sheets
        for sec in s["sections"]
        for sp in [sec["subProcedures"]]
    )
    print(f"  Total sub-procedures: {total_subs}")

    print("Building workbook...")
    wb = Workbook()

    # Sheet 1: Methodology
    build_methodology_sheet(wb)
    print("  [1/7] Methodology & Approach")

    # Sheets 2-6: Assessment sheets
    all_control_scores = []
    for sheet_data in sheets:
        control_scores = build_assessment_sheet(wb, sheet_data)
        all_control_scores.extend(control_scores)
        n = sheet_data["sheetNumber"]
        name = sheet_data["sheetName"]
        subs = sum(len(sec["subProcedures"]) for sec in sheet_data["sections"])
        controls = len(sheet_data["sections"])
        print(f"  [{n}/7] {name} ({controls} controls, {subs} sub-procedures)")

    # Sheet 7: Scoring Summary (at the end)
    build_scoring_sheet(wb, all_control_scores)
    print(f"  [7/7] Scoring Summary ({len(all_control_scores)} controls scored)")

    # Save
    wb.save(str(OUTPUT_FILE))
    print(f"\nSaved: {OUTPUT_FILE}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    print(f"  Sub-procedures: {total_subs}")
    print(f"  Controls with scoring: {len(all_control_scores)}")


if __name__ == "__main__":
    main()
